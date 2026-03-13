"""
Website service for managing lab website PRs.

Handles:
- Reading/writing people.xlsx (members and alumni sheets)
- Uploading photos to images/people/
- Creating branches and PRs for website changes
"""

import base64
import io
import logging
import re
from dataclasses import dataclass
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Optional, Tuple

import openpyxl
from github import Github, GithubException
from github.Repository import Repository

logger = logging.getLogger(__name__)


class MemberRole(Enum):
    """Valid roles for lab members."""
    GRAD_STUDENT = "Graduate Student"
    UNDERGRAD = "Undergraduate"
    POSTDOC = "Postdoctoral Researcher"
    LAB_MANAGER = "Lab Manager"
    RESEARCH_SCIENTIST = "Research Scientist"


class GradType(Enum):
    """Graduate student types."""
    DOCTORAL = "Doctoral"
    MASTERS = "Masters"


@dataclass
class WebsiteContent:
    """Content to be added/updated on the website."""
    name: str
    name_url: Optional[str] = None
    role: str = ""
    bio: str = ""
    links_html: str = ""
    image_filename: str = ""
    image_data: Optional[bytes] = None


@dataclass
class AlumniContent:
    """Content for alumni entry."""
    name: str
    name_url: Optional[str] = None
    years: str = ""
    current_position: str = ""
    current_position_url: Optional[str] = None


class WebsiteService:
    """Service for managing contextlab.github.io website content."""

    WEBSITE_REPO = "ContextLab/contextlab.github.io"
    PEOPLE_FILE = "data/people.xlsx"
    IMAGES_PATH = "images/people"
    CV_FILE = "documents/JRM_CV.tex"
    MEMBERS_SHEET = "members"

    # Column indices in members sheet (0-indexed)
    MEMBER_COLS = {
        "image": 0,
        "name": 1,
        "name_url": 2,
        "role": 3,
        "bio": 4,
        "links_html": 5,
    }

    # Column indices in alumni sheets
    ALUMNI_COLS = {
        "name": 0,
        "name_url": 1,
        "years": 2,
        "current_position": 3,
        "current_position_url": 4,
    }

    # Map roles to alumni sheets
    ALUMNI_SHEET_MAP = {
        MemberRole.GRAD_STUDENT: "alumni_grads",
        MemberRole.UNDERGRAD: "alumni_undergrads",
        MemberRole.POSTDOC: "alumni_postdocs",
        MemberRole.LAB_MANAGER: "alumni_managers",
        MemberRole.RESEARCH_SCIENTIST: "alumni_managers",
    }

    def __init__(self, token: str):
        """Initialize website service with GitHub token."""
        self.github = Github(token)
        self._repo: Optional[Repository] = None

    @property
    def repo(self) -> Repository:
        """Get the website repository (lazy loaded)."""
        if self._repo is None:
            self._repo = self.github.get_repo(self.WEBSITE_REPO)
        return self._repo

    def get_people_xlsx(self, ref: str = "main") -> openpyxl.Workbook:
        """Download and parse people.xlsx from the repository."""
        contents = self.repo.get_contents(self.PEOPLE_FILE, ref=ref)
        xlsx_data = base64.b64decode(contents.content)
        return openpyxl.load_workbook(io.BytesIO(xlsx_data))

    def get_current_members(self) -> list[dict]:
        """Get list of current members from the spreadsheet."""
        wb = self.get_people_xlsx()
        ws = wb[self.MEMBERS_SHEET]
        members = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[self.MEMBER_COLS["name"]]
            if name:
                members.append({
                    "image": row[self.MEMBER_COLS["image"]] or "",
                    "name": name,
                    "name_url": row[self.MEMBER_COLS["name_url"]] or "",
                    "role": row[self.MEMBER_COLS["role"]] or "",
                    "bio": row[self.MEMBER_COLS["bio"]] or "",
                    "links_html": row[self.MEMBER_COLS["links_html"]] or "",
                })
        return members

    def get_alumni_sheets(self) -> list[str]:
        """Get list of alumni sheet names."""
        wb = self.get_people_xlsx()
        return [name for name in wb.sheetnames if name.startswith("alumni_")]

    def find_member_by_name(self, name: str) -> Optional[Tuple[int, dict]]:
        """Find a member by name and return (row_index, data)."""
        wb = self.get_people_xlsx()
        ws = wb[self.MEMBERS_SHEET]

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            member_name = row[self.MEMBER_COLS["name"]]
            if member_name and member_name.lower() == name.lower():
                return idx, {
                    "image": row[self.MEMBER_COLS["image"]] or "",
                    "name": member_name,
                    "name_url": row[self.MEMBER_COLS["name_url"]] or "",
                    "role": row[self.MEMBER_COLS["role"]] or "",
                    "bio": row[self.MEMBER_COLS["bio"]] or "",
                    "links_html": row[self.MEMBER_COLS["links_html"]] or "",
                }
        return None

    def generate_image_filename(self, name: str) -> str:
        """Generate image filename from member name."""
        # Convert "First Last" to "first_last.png"
        # Handle multiple spaces and special characters
        clean_name = re.sub(r"[^\w\s]", "", name)
        parts = clean_name.lower().split()
        return "_".join(parts) + ".png"

    def create_onboarding_pr(
        self,
        content: WebsiteContent,
        cv_entry: Optional[str] = None,
        cv_section: Optional[str] = None,
        slack_user_id: str = "",
    ) -> Tuple[bool, Optional[str], Optional[str]]:
        """
        Create a PR to add a new member to the website.

        Args:
            content: Website content for the new member
            cv_entry: LaTeX entry to add to CV (e.g., "\\item Name (Doctoral student; 2025 -- )")
            cv_section: CV section to add to ("Postdoctoral Advisees", "Graduate Advisees", "Undergraduate Advisees")
            slack_user_id: Slack user ID for branch naming

        Returns:
            Tuple of (success, pr_url or error_message, branch_name)
        """
        branch_name = f"onboarding/{slack_user_id}-{content.name.replace(' ', '-').lower()}"

        try:
            # Get default branch ref
            main_ref = self.repo.get_git_ref("heads/main")
            main_sha = main_ref.object.sha

            # Create new branch
            try:
                self.repo.create_git_ref(f"refs/heads/{branch_name}", main_sha)
            except GithubException as e:
                if e.status == 422:  # Branch already exists
                    self.repo.get_git_ref(f"heads/{branch_name}").delete()
                    self.repo.create_git_ref(f"refs/heads/{branch_name}", main_sha)
                else:
                    raise

            # 1. Upload image if provided
            if content.image_data and content.image_filename:
                image_path = f"{self.IMAGES_PATH}/{content.image_filename}"
                try:
                    # Check if file exists
                    existing = self.repo.get_contents(image_path, ref=branch_name)
                    self.repo.update_file(
                        path=image_path,
                        message=f"Update photo for {content.name}",
                        content=content.image_data,
                        sha=existing.sha,
                        branch=branch_name,
                    )
                except GithubException:
                    # File doesn't exist, create it
                    self.repo.create_file(
                        path=image_path,
                        message=f"Add photo for {content.name}",
                        content=content.image_data,
                        branch=branch_name,
                    )

            # 2. Update people.xlsx
            wb = self.get_people_xlsx(ref=branch_name)
            ws = wb[self.MEMBERS_SHEET]

            # Check if member already exists (idempotency)
            existing_row = None
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                member_name = row[self.MEMBER_COLS["name"]]
                if member_name and member_name.lower() == content.name.lower():
                    existing_row = idx
                    logger.info(f"Member {content.name} already exists at row {idx}, updating")
                    break

            if existing_row:
                # Update existing row
                target_row = existing_row
            else:
                # Add new row
                target_row = ws.max_row + 1

            ws.cell(row=target_row, column=1, value=content.image_filename)
            ws.cell(row=target_row, column=2, value=content.name)
            ws.cell(row=target_row, column=3, value=content.name_url or None)
            ws.cell(row=target_row, column=4, value=content.role)
            ws.cell(row=target_row, column=5, value=content.bio)
            ws.cell(row=target_row, column=6, value=content.links_html or None)

            # Save to bytes
            xlsx_bytes = io.BytesIO()
            wb.save(xlsx_bytes)
            xlsx_content = xlsx_bytes.getvalue()

            # Get current file SHA and update
            people_file = self.repo.get_contents(self.PEOPLE_FILE, ref=branch_name)
            self.repo.update_file(
                path=self.PEOPLE_FILE,
                message=f"Add {content.name} to members",
                content=xlsx_content,
                sha=people_file.sha,
                branch=branch_name,
            )

            # 3. Update CV if entry provided
            if cv_entry and cv_section:
                self._add_cv_entry(branch_name, cv_entry, cv_section, content.name)

            # 4. Create PR
            pr_body = self._build_onboarding_pr_body(content, cv_entry)
            pr = self.repo.create_pull(
                title=f"Add {content.name} to lab members",
                body=pr_body,
                head=branch_name,
                base="main",
            )

            logger.info(f"Created website PR: {pr.html_url}")
            return True, pr.html_url, branch_name

        except GithubException as e:
            error_msg = f"Error creating website PR: {e}"
            logger.error(error_msg)
            # Cleanup branch on failure
            try:
                self.repo.get_git_ref(f"heads/{branch_name}").delete()
            except Exception:
                pass
            return False, error_msg, None

    def create_offboarding_pr(
        self,
        member_name: str,
        alumni_content: AlumniContent,
        alumni_sheet: str,
        cv_update: Optional[Tuple[str, str]] = None,
        slack_user_id: str = "",
    ) -> Tuple[bool, Optional[str], Optional[str]]:
        """
        Create a PR to move a member to alumni.

        Args:
            member_name: Current name in members sheet
            alumni_content: Content for alumni entry
            alumni_sheet: Target alumni sheet (e.g., "alumni_grads")
            cv_update: Tuple of (old_entry_pattern, new_entry) for CV update
            slack_user_id: Slack user ID for branch naming

        Returns:
            Tuple of (success, pr_url or error_message, branch_name)
        """
        branch_name = f"offboarding/{slack_user_id}-{member_name.replace(' ', '-').lower()}"

        try:
            # Get default branch ref
            main_ref = self.repo.get_git_ref("heads/main")
            main_sha = main_ref.object.sha

            # Create new branch
            try:
                self.repo.create_git_ref(f"refs/heads/{branch_name}", main_sha)
            except GithubException as e:
                if e.status == 422:
                    self.repo.get_git_ref(f"heads/{branch_name}").delete()
                    self.repo.create_git_ref(f"refs/heads/{branch_name}", main_sha)
                else:
                    raise

            # Update people.xlsx
            wb = self.get_people_xlsx(ref=branch_name)
            members_ws = wb[self.MEMBERS_SHEET]

            # Find and remove from members
            member_row = None
            for idx, row in enumerate(members_ws.iter_rows(min_row=2, values_only=True), start=2):
                name = row[self.MEMBER_COLS["name"]]
                if name and name.lower() == member_name.lower():
                    member_row = idx
                    break

            if member_row:
                members_ws.delete_rows(member_row)

            # Add to alumni sheet
            if alumni_sheet not in wb.sheetnames:
                logger.warning(f"Alumni sheet {alumni_sheet} not found, creating it")
                wb.create_sheet(alumni_sheet)

            alumni_ws = wb[alumni_sheet]

            # Check if alumni entry already exists (idempotency)
            existing_alumni_row = None
            for idx, row in enumerate(alumni_ws.iter_rows(min_row=2, values_only=True), start=2):
                alumni_name = row[0]  # Name is always first column
                if alumni_name and alumni_name.lower() == alumni_content.name.lower():
                    existing_alumni_row = idx
                    logger.info(f"Alumni {alumni_content.name} already exists at row {idx}, updating")
                    break

            if existing_alumni_row:
                target_row = existing_alumni_row
            else:
                target_row = alumni_ws.max_row + 1

            # Alumni sheets have different columns for undergrads
            if alumni_sheet == "alumni_undergrads":
                # Undergrads only have: name, years
                alumni_ws.cell(row=target_row, column=1, value=alumni_content.name)
                alumni_ws.cell(row=target_row, column=2, value=alumni_content.years)
            else:
                alumni_ws.cell(row=target_row, column=1, value=alumni_content.name)
                alumni_ws.cell(row=target_row, column=2, value=alumni_content.name_url or None)
                alumni_ws.cell(row=target_row, column=3, value=alumni_content.years)
                alumni_ws.cell(row=target_row, column=4, value=alumni_content.current_position)
                alumni_ws.cell(row=target_row, column=5, value=alumni_content.current_position_url or None)

            # Save to bytes
            xlsx_bytes = io.BytesIO()
            wb.save(xlsx_bytes)
            xlsx_content = xlsx_bytes.getvalue()

            # Update file
            people_file = self.repo.get_contents(self.PEOPLE_FILE, ref=branch_name)
            self.repo.update_file(
                path=self.PEOPLE_FILE,
                message=f"Move {member_name} to alumni",
                content=xlsx_content,
                sha=people_file.sha,
                branch=branch_name,
            )

            # Update CV if provided
            if cv_update:
                old_pattern, new_entry = cv_update
                self._update_cv_entry(branch_name, old_pattern, new_entry, member_name)

            # Create PR
            pr_body = self._build_offboarding_pr_body(member_name, alumni_content, cv_update)
            pr = self.repo.create_pull(
                title=f"Move {member_name} to alumni ({alumni_sheet})",
                body=pr_body,
                head=branch_name,
                base="main",
            )

            logger.info(f"Created offboarding PR: {pr.html_url}")
            return True, pr.html_url, branch_name

        except GithubException as e:
            error_msg = f"Error creating offboarding PR: {e}"
            logger.error(error_msg)
            try:
                self.repo.get_git_ref(f"heads/{branch_name}").delete()
            except Exception:
                pass
            return False, error_msg, None

    def _cv_entry_exists(self, cv_content: str, name: str, section: str) -> bool:
        """Check if a CV entry already exists for this person in the section."""
        # Look for \item Name in the section
        section_marker = f"\\textit{{{section}}}:"
        if section_marker not in cv_content:
            return False

        section_idx = cv_content.index(section_marker)

        # Find the end of this section (next \textit or \end{enumerate})
        next_section = cv_content.find("\\textit{", section_idx + 1)
        end_enumerate = cv_content.find("\\end{etaremune}", section_idx)

        if next_section == -1:
            section_end = end_enumerate
        elif end_enumerate == -1:
            section_end = next_section
        else:
            section_end = min(next_section, end_enumerate)

        section_content = cv_content[section_idx:section_end]

        # Check for \item Name (case-insensitive, handles variations like middle names)
        # Match pattern: \item FirstName ... LastName
        name_parts = name.split()
        if len(name_parts) >= 2:
            first_name = name_parts[0]
            last_name = name_parts[-1]
            pattern = rf"\\item\s+{re.escape(first_name)}.*{re.escape(last_name)}"
            if re.search(pattern, section_content, re.IGNORECASE):
                return True

        return False

    def _add_cv_entry(self, branch: str, entry: str, section: str, name: str):
        """Add an entry to the CV LaTeX file."""
        try:
            cv_file = self.repo.get_contents(self.CV_FILE, ref=branch)
            cv_content = base64.b64decode(cv_file.content).decode("utf-8")

            # Check if entry already exists
            if self._cv_entry_exists(cv_content, name, section):
                logger.info(f"CV entry for {name} already exists in {section}, skipping")
                return

            # Find the section and add entry at top of list
            section_marker = f"\\textit{{{section}}}:"
            if section_marker not in cv_content:
                logger.warning(f"CV section '{section}' not found")
                return

            # Find \begin{etaremune} after the section marker
            section_idx = cv_content.index(section_marker)
            begin_idx = cv_content.index("\\begin{etaremune}", section_idx)
            insert_idx = cv_content.index("\n", begin_idx) + 1

            # Insert the new entry
            new_content = cv_content[:insert_idx] + entry + "\n" + cv_content[insert_idx:]

            self.repo.update_file(
                path=self.CV_FILE,
                message=f"Add {name} to CV mentorship section",
                content=new_content.encode("utf-8"),
                sha=cv_file.sha,
                branch=branch,
            )
        except Exception as e:
            logger.error(f"Error adding CV entry: {e}")

    def _update_cv_entry(self, branch: str, old_pattern: str, new_entry: str, name: str):
        """Update an existing entry in the CV LaTeX file."""
        try:
            cv_file = self.repo.get_contents(self.CV_FILE, ref=branch)
            cv_content = base64.b64decode(cv_file.content).decode("utf-8")

            # Replace the old entry with the new one
            if old_pattern in cv_content:
                new_content = cv_content.replace(old_pattern, new_entry)
                self.repo.update_file(
                    path=self.CV_FILE,
                    message=f"Update {name} CV entry for alumni transition",
                    content=new_content.encode("utf-8"),
                    sha=cv_file.sha,
                    branch=branch,
                )
            else:
                logger.warning(f"CV entry pattern not found: {old_pattern}")
        except Exception as e:
            logger.error(f"Error updating CV entry: {e}")

    def delete_branch(self, branch_name: str) -> bool:
        """Delete a branch (for cleanup after PR merge/close)."""
        try:
            ref = self.repo.get_git_ref(f"heads/{branch_name}")
            ref.delete()
            return True
        except GithubException:
            return False

    def _build_onboarding_pr_body(self, content: WebsiteContent, cv_entry: Optional[str]) -> str:
        """Build PR description for onboarding."""
        body = f"""## New Lab Member: {content.name}

**Role:** {content.role}

**Bio:**
> {content.bio}

**Website:** {content.name_url or 'None'}

---

### Changes
- [ ] Photo uploaded to `images/people/{content.image_filename}`
- [ ] Member added to `people.xlsx` members sheet
"""
        if cv_entry:
            body += f"- [ ] CV updated with new mentee entry\n"

        body += """
After merging, GitHub Actions will automatically rebuild the website.

---
Generated by CDL Onboarding Bot
"""
        return body

    def _build_offboarding_pr_body(
        self,
        name: str,
        alumni: AlumniContent,
        cv_update: Optional[Tuple[str, str]]
    ) -> str:
        """Build PR description for offboarding."""
        body = f"""## Transition to Alumni: {name}

**Years Active:** {alumni.years}
**Current Position:** {alumni.current_position}
**Position URL:** {alumni.current_position_url or 'None'}

---

### Changes
- [ ] Removed from `members` sheet
- [ ] Added to alumni sheet
"""
        if cv_update:
            body += "- [ ] CV entry updated with end date and current position\n"

        body += """
After merging, GitHub Actions will automatically rebuild the website.

---
Generated by CDL Onboarding Bot
"""
        return body


def build_cv_entry(
    name: str,
    role: MemberRole,
    grad_type: Optional[GradType] = None,
    grad_field: Optional[str] = None,
    year: Optional[int] = None,
) -> Tuple[Optional[str], Optional[str]]:
    """
    Build a CV entry for a new member.

    Returns:
        Tuple of (entry_text, section_name) or (None, None) if role doesn't get CV entry
    """
    if year is None:
        year = datetime.now().year

    if role == MemberRole.LAB_MANAGER or role == MemberRole.RESEARCH_SCIENTIST:
        return None, None

    if role == MemberRole.POSTDOC:
        entry = f"\\item {name} ({year} -- )"
        section = "Postdoctoral Advisees"
    elif role == MemberRole.UNDERGRAD:
        entry = f"\\item {name} ({year} -- )"
        section = "Undergraduate Advisees"
    elif role == MemberRole.GRAD_STUDENT:
        if grad_type == GradType.MASTERS and grad_field:
            entry = f"\\item {name} (Masters student, {grad_field}; {year} -- )"
        elif grad_type == GradType.MASTERS:
            entry = f"\\item {name} (Masters student; {year} -- )"
        else:
            entry = f"\\item {name} (Doctoral student; {year} -- )"
        section = "Graduate Advisees"
    else:
        return None, None

    return entry, section


def build_cv_update_for_offboarding(
    name: str,
    role: MemberRole,
    start_year: int,
    end_year: int,
    current_position: str,
    grad_type: Optional[GradType] = None,
    grad_field: Optional[str] = None,
) -> Optional[Tuple[str, str]]:
    """
    Build CV update pattern for offboarding.

    Returns:
        Tuple of (old_entry_pattern, new_entry) or None if role doesn't have CV entry
    """
    if role == MemberRole.LAB_MANAGER or role == MemberRole.RESEARCH_SCIENTIST:
        return None

    if role == MemberRole.POSTDOC:
        old = f"\\item {name} ({start_year} -- )"
        new = f"\\item {name} ({start_year} -- {end_year}; current position: {current_position})"
    elif role == MemberRole.UNDERGRAD:
        old = f"\\item {name} ({start_year} -- )"
        new = f"\\item {name} ({start_year} -- {end_year})"
    elif role == MemberRole.GRAD_STUDENT:
        if grad_type == GradType.MASTERS and grad_field:
            old = f"\\item {name} (Masters student, {grad_field}; {start_year} -- )"
            new = f"\\item {name} (Masters student, {grad_field}; {start_year} -- {end_year}; current position: {current_position})"
        elif grad_type == GradType.MASTERS:
            old = f"\\item {name} (Masters student; {start_year} -- )"
            new = f"\\item {name} (Masters student; {start_year} -- {end_year}; current position: {current_position})"
        else:
            old = f"\\item {name} (Doctoral student; {start_year} -- )"
            new = f"\\item {name} (Doctoral student; {start_year} -- {end_year}; current position: {current_position})"
    else:
        return None

    return old, new
